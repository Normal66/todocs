# By Constantine Sidorov, 2026

from __future__ import annotations

import re
from dataclasses import dataclass
from pathlib import Path

import openpyxl


def _slug_header(s: str) -> str:
    s = s.strip().replace("\n", " ")
    s = re.sub(r"\s+", " ", s)
    s = re.sub(r"[^\w]+", "_", s, flags=re.UNICODE)
    s = s.strip("_")
    return s or "COL"


@dataclass(frozen=True)
class HeaderMap:
    by_col: dict[str, str]  # A..I -> slug
    by_col_raw: dict[str, str]  # A..I -> raw header text

    def _find_by_any_hint(self, *hints: str) -> str | None:
        hints_l = [h.lower() for h in hints]
        for col, raw in self.by_col_raw.items():
            raw_l = raw.lower()
            if any(h in raw_l for h in hints_l):
                return self.by_col.get(col)
        return None

    def _find_by_all_hints(self, *hints: str) -> str | None:
        hints_l = [h.lower() for h in hints]
        for col, raw in self.by_col_raw.items():
            raw_l = raw.lower()
            if all(h in raw_l for h in hints_l):
                return self.by_col.get(col)
        return None

    @property
    def inv_no(self) -> str:  # E
        return self._find_by_any_hint("инв") or self.by_col.get("E", "INV_NO")

    @property
    def name(self) -> str:  # B
        return self._find_by_any_hint("наимен") or self.by_col.get("B", "NAME")

    @property
    def serial(self) -> str:  # F
        return self._find_by_any_hint("серийн", "завод") or self.by_col.get("F", "SERIAL")

    @property
    def defect_mon(self) -> str:  # H
        return (
            self._find_by_all_hints("неисп", "монитор")
            or self._find_by_any_hint("неисп", "монитор")
            or self._find_by_any_hint("монитор")
            or self.by_col.get("H", "DEFECT_MON")
        )

    @property
    def defect_sys(self) -> str:  # I
        return (
            self._find_by_all_hints("неисп", "сист")
            or self._find_by_any_hint("неисп", "сист")
            or self._find_by_any_hint("систем")
            or self.by_col.get("I", "DEFECT_SYS")
        )


def read_headers(xlsx_path: Path) -> HeaderMap:
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    # Берём первые 9 колонок (A..I) как в исходном Excel.
    cols = list("ABCDEFGHI")
    by_col: dict[str, str] = {}
    by_col_raw: dict[str, str] = {}
    for col, v in zip(cols, row[: len(cols)], strict=False):
        if v is None:
            continue
        raw = str(v)
        by_col_raw[col] = raw
        by_col[col] = _slug_header(raw)
    return HeaderMap(by_col=by_col, by_col_raw=by_col_raw)


def main() -> None:
    repo_root = Path(__file__).resolve().parents[1]
    xlsx = next(repo_root.glob("*.xlsx"))
    h = read_headers(xlsx)
    print("xlsx:", xlsx.name)
    print("headers:", h.by_col)
    print("inv_no:", h.inv_no)
    print("name:", h.name)
    print("serial:", h.serial)
    print("def_mon:", h.defect_mon)
    print("def_sys:", h.defect_sys)


if __name__ == "__main__":
    main()

